"""生成BOSS数据汇总Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 样式定义 =====
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 费用级别颜色
fee_fills = {
    "5万": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "10万": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "15万": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
    "20万": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
    "30万": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
    "100万": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}

HEADERS = [
    "序号", "中文名称", "英文名称", "怪物ID", "等级",
    "所在地图", "传送费用(金币)", "数据来源", "备注"
]
COL_WIDTHS = [6, 22, 22, 12, 8, 14, 18, 28, 30]


def write_sheet(ws, title, data):
    """写入一个工作表"""
    ws.title = title

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # 数据
    for ri, row in enumerate(data, 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.alignment = left_align if ci in (2, 3, 8, 9) else data_align
            cell.border = thin_border

            # 按费用着色
            fee_str = row[6]  # 传送费用列
            for key, fill in fee_fills.items():
                if key in str(fee_str):
                    cell.fill = fill
                    break

    # 列宽
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 冻结表头
    ws.freeze_panes = "A3"
    # 自动筛选
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{2 + len(data)}"


# =============================================
# Sheet 1: 远征BOSS
# =============================================
expedition_data = [
    [1, "鱼王", "Pianus", 8510000, 110, 230040420, 100000, "万能传送.js / Mob.wz / 卡片收集.js", "左侧鱼王为精英BOSS(8520000)"],
    [2, "闹钟", "Papulatus", 8500002, 125, 220080000, 100000, "万能传送.js / PapulatusBattle.js / Mob.wz", "本体ID:8500002; 座钟:8500001"],
    [3, "扎昆", "Zakum", 8800002, 140, 211042300, 300000, "万能传送.js / ZakumBattle.js / Mob.wz", "本体ID:8800002; 手臂:8800003~8800010"],
    [4, "妖僧", "Yao Seng", 9600025, 0, 702070400, 300000, "万能传送.js / YaoSengPQ.js / String.wz", "武林妖僧; PQ副本BOSS; 等级待确认"],
    [5, "暗黑龙王", "Horntail", 8810018, 160, 240050400, 1000000, "万能传送.js / HorntailBattle.js / Mob.wz", "灵魂ID:8810018; 头部:8810000; 三龙合一"],
    [6, "品克缤", "Pink Bean", 8820001, 180, 270050000, 1000000, "万能传送.js / PinkBeanBattle.js / Mob.wz", "时间的宠儿-品克缤; 带5个女神雕像"],
]

# =============================================
# Sheet 2: 精英BOSS
# =============================================
elite_data = [
    # 低级 5万
    [1, "蜗牛王", "Mano", 2220000, 20, 104000400, 50000, "万能传送.js / AreaBossMano.js / Mob.wz", ""],
    [2, "树妖王", "Stumpy", 3220000, 35, 101030404, 50000, "万能传送.js / AreaBossStumpy.js / Mob.wz", ""],
    [3, "蘑菇王", "Mushmom", 6130101, 60, 100000005, 50000, "万能传送.js / 卡片收集.js / Mob.wz", ""],
    # 中低级 10万
    [4, "巨型蝙蝠怪", "Balrog", 8830000, 105, 105100100, 100000, "万能传送.js / 卡片收集.js / Mob.wz", "远征级蝙蝠怪; 另有Easy(8830010)/Hard(8830003)模式"],
    [5, "左鱼王", "Left Pianus", 8520000, 110, 230040410, 100000, "万能传送.js / 卡片收集.js / Mob.wz", "鱼王(8510000)的左侧变体"],
    [6, "大宇", "Deo", 3220001, 38, 260010201, 100000, "万能传送.js / AreaBossDeo.js / Mob.wz", "阿里安特地区"],
    [7, "海龙王", "King Clang", 5220001, 55, 110040000, 100000, "万能传送.js / AreaBossKingClang.js / Mob.wz", "String.wz名: 巨居蟹; 佛罗里达海滩"],
    [8, "浮士德", "Faust", 5220002, 50, 100040105, 100000, "万能传送.js / AreaBossFaust1.js / Mob.wz", "另有AreaBossFaust2.js(地图100040106)"],
    [9, "僵尸蘑菇王", "Zombie Mushmom", 6300005, 60, 105070002, 100000, "万能传送.js / String.wz / Mob.wz", "蚂蚁洞广场"],
    [10, "小巴尔罗格", "Jr. Balrog", 8130100, 80, 105090900, 100000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 蝙蝠怪; 龙穴入口"],
    [11, "时间门神", "Timer", 5220003, 59, 220050100, 100000, "万能传送.js / AreaBossTimer2.js / Mob.wz", "String.wz名: 提莫; 另有Timer1(220050000)/Timer3(220050200)"],
    [12, "瑟鲁夫", "Seruf", 4220001, 45, 230020100, 100000, "万能传送.js / AreaBossSeruf.js / Mob.wz", "水下世界-危海峡谷"],
    [13, "暗影", "Shade", 5090000, 56, 103000105, 100000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 谢尔德; 废弃都市地铁"],
    # 中级 15万
    [14, "蜈蚣王", "Centipede", 5220004, 50, 251010102, 150000, "万能传送.js / AreaBossCentipede.js / Mob.wz", "String.wz名: 巨型蜈蚣; 百草堂"],
    [15, "迪勒", "Dyle", 6220000, 65, 107000300, 150000, "万能传送.js / AreaBossDyle.js / Mob.wz", "String.wz名: 多尔; 林中之城沼泽"],
    [16, "雪女", "Snow Witch", 6090001, 64, 211010000, 150000, "万能传送.js / Mob.wz", "String.wz名: 雪山魔女; 冰峰雪域"],
    [17, "书生幽灵", "Ghost Scholar", 6090003, 62, 222010300, 150000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 书生鬼; 童话村狐狸山坡"],
    [18, "艾利杰", "Eliza", 8220000, 83, 200010300, 150000, "万能传送.js / AreaBossEliza1.js / Mob.wz", "天空之城"],
    [19, "利奇", "Riche", 6090000, 74, 211041100, 150000, "万能传送.js / Reactor2119003.js / Mob.wz", "String.wz名: 黑山老妖; 废矿区"],
    [20, "鲁鲁莫", "Rurumo", 6090004, 63, 261020200, 150000, "万能传送.js / Reactor2619005.js / Mob.wz", "String.wz名: 陆陆猫; 玛加提亚"],
    [21, "芝诺", "Zeno", 6220001, 65, 221040301, 150000, "万能传送.js / AreaBossZeno.js / Mob.wz", "String.wz名: 朱诺; 地球防御本部"],
    [22, "竹武士", "Bamboo Warrior", 6090002, 68, 800020120, 150000, "万能传送.js / AreaBossBamboo.js / Mob.wz", "String.wz名: 青竹武士; 古代神社"],
    # 中高级 20万
    [23, "无头骑士", "Headless Horseman", 9400571, 50, 682000001, 200000, "万能传送.js / String.wz / Mob.wz", "万圣节地图; 特殊BOSS"],
    [24, "雪男", "Snowman", 8220001, 90, 211040101, 200000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 驮狼雪人; 冰峰雪域深处"],
    [25, "九尾狐", "Nine-Tailed Fox", 7220001, 70, 800020130, 200000, "万能传送.js / AreaBossNineTailedFox.js / Mob.wz", "古代神社"],
    [26, "泰伦", "Tae Roon", 7220000, 71, 250010304, 200000, "万能传送.js / AreaBossTaeRoon.js / Mob.wz", "String.wz名: 肯德熊; 武陵"],
    [27, "智慧猫王", "King Sage Cat", 7220002, 77, 250010504, 200000, "万能传送.js / AreaBossKingSageCat.js / Mob.wz", "String.wz名: 妖怪禅师; 武陵"],
    [28, "奇美拉", "Chimera", 8220002, 85, 261030000, 200000, "万能传送.js / AreaBossKimera.js / Mob.wz", "String.wz名: 吉米拉; 玛加提亚"],
    [29, "零食吧", "Snack Bar", 8220009, 85, 105090310, 200000, "万能传送.js / AreaBossSnackBar.js / Mob.wz", "String.wz名: 小吃店; 龙穴区域"],
    # 高级 30万
    [30, "喷火龙", "Manon", 8180000, 105, 240020401, 300000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 火焰龙; 神木村"],
    [31, "格瑞芬多", "Griffey", 8180001, 105, 240020101, 300000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 天鹰; 神木村"],
    [32, "船老大", "Captain", 8150000, 100, 251010404, 300000, "万能传送.js / MonsterBook.img.xml / Mob.wz", "String.wz名: 蝙蝠魔; 海盗团老巢"],
    [33, "利维坦", "Leviathan", 8220003, 120, 240040401, 300000, "万能传送.js / AreaBossLeviathan.js / Mob.wz", "String.wz名: 大海兽; 神木村-龙之巢"],
    [34, "多多", "Dodo", 8220004, 105, 270010500, 300000, "万能传送.js / Quest.wz / Mob.wz", "时间神殿-追忆之路; 时间鲸鱼"],
    [35, "莉里诺斯", "Lilynouch", 8220005, 105, 270020500, 300000, "万能传送.js / Quest.wz / Mob.wz", "String.wz名: 玄冰独角兽; 时间神殿-后悔之路"],
    [36, "雷卡", "Lyka", 8220006, 110, 270030500, 300000, "万能传送.js / Mob.wz", "时间神殿-忘却之路"],
]

# =============================================
# Sheet 3: 特殊BOSS (PQ/副本/活动)
# =============================================
special_data = [
    [1, "巴尔罗格(Hard)", "Balrog (Hard)", 8830003, 105, 105100300, 0, "BalrogBattle.js", "远征挑战模式"],
    [2, "巴尔罗格(Easy)", "Balrog (Easy)", 8830010, 105, 105100400, 0, "BalrogBattle_Easy.js", "远征简单模式"],
    [3, "印第安老斑鸠", "Custom Boss 1", 9400609, 0, 677000005, 0, "AreaBossDoor2.js", "门BOSS系列"],
    [4, "黑暗独角兽", "Custom Boss 2", 9400610, 0, 677000003, 0, "AreaBossDoor1.js", "门BOSS系列"],
    [5, "雪之猫女", "Custom Boss 3", 9400611, 0, 677000007, 0, "AreaBossDoor6.js", "门BOSS系列"],
    [6, "牛魔王", "Custom Boss 4", 9400612, 0, 677000001, 0, "AreaBossDoor5.js", "门BOSS系列"],
    [7, "沃勒福", "Custom Boss 5", 9400613, 0, 677000009, 0, "AreaBossDoor3.js", "门BOSS系列"],
    [8, "牛魔王(2)", "Custom Boss 6", 9400633, 0, 677000012, 0, "AreaBossDoor4.js", "门BOSS系列"],
    [9, "武功", "WuGong PQ Boss", 9600009, 50, 0, 0, "WuGongPQ.js", "武功PQ副本BOSS"],
]

# =============================================
# 写入工作表
# =============================================
ws1 = wb.active
write_sheet(ws1, "远征BOSS", expedition_data)

ws2 = wb.create_sheet()
write_sheet(ws2, "精英BOSS", elite_data)

ws3 = wb.create_sheet()
write_sheet(ws3, "特殊BOSS", special_data)

# =============================================
# Sheet 4: 汇总说明
# =============================================
ws4 = wb.create_sheet("数据说明")
notes = [
    ["数据来源"],
    ["1. 万能传送.js — gms-server/scripts-zh-CN/BeiDouSpecial/万能传送.js (BOSS列表和传送地图)"],
    ["2. Mob.wz — gms-server/wz/Mob.wz/ (怪物等级、属性等基础数据)"],
    ["3. String.wz/Mob.img.xml — 怪物中文/英文名称"],
    ["4. String.wz/MonsterBook.img.xml — 怪物所属地图、掉落奖励"],
    ["5. AreaBoss*.js — gms-server/scripts-zh-CN/event/ (野外BOSS刷新脚本)"],
    ["6. *Battle.js — 远征BOSS战斗脚本 (ZakumBattle, HorntailBattle, PinkBeanBattle等)"],
    ["7. 卡片收集.js — gms-server/scripts-zh-CN/BeiDouSpecial/ (怪物卡片与BOSS对照)"],
    ["8. Reactor*.js — 反应堆脚本 (部分BOSS通过反应堆触发)"],
    ["9. Quest.wz — 任务文本中提及的BOSS信息"],
    [""],
    ["数据说明"],
    ["- '等级' 取自 Mob.wz XML 中第一个 <int name=\"level\"> 值，表示怪物基础等级"],
    ["- BOSS实际强度受服务器配置(exp/HP倍率等)影响，可能高于基础等级"],
    ["- '0' 表示该数据暂未从WZ文件中确认"],
    ["- 精英BOSS按传送费用分为五档: 5万(低级) → 10万(中低级) → 15万(中级) → 20万(中高级) → 30万(高级)"],
    ["- 远征BOSS的传送地图为BOSS入口地图，实际战斗房间可能不同"],
    ["- 部分BOSS存在多种形态/阶段(如扎昆手臂、暗黑龙王头部/翅膀/尾巴等)，表中仅列本体ID"],
    [""],
    ["更新记录"],
    ["2026-05-25: 初始版本，收录远征BOSS 6个、精英BOSS 36个、特殊BOSS 9个"],
    ["数据整理: 北斗项目组"],
]

for ri, row in enumerate(notes, 1):
    for ci, val in enumerate(row, 1):
        cell = ws4.cell(row=ri, column=ci, value=val)
        if ri == 1 or (len(notes) > ri and notes[ri] and notes[ri][0].startswith("数据")):
            cell.font = Font(name="微软雅黑", size=11, bold=True)
        else:
            cell.font = Font(name="微软雅黑", size=10)

ws4.column_dimensions["A"].width = 120

output_path = "E:/pro/BeiDou-Server_xy/resource_doc/BOSS数据汇总.xlsx"
wb.save(output_path)
print(f"Excel已生成: {output_path}")
print(f"远征BOSS: {len(expedition_data)} 个")
print(f"精英BOSS: {len(elite_data)} 个")
print(f"特殊BOSS: {len(special_data)} 个")
print(f"合计: {len(expedition_data) + len(elite_data) + len(special_data)} 个")
