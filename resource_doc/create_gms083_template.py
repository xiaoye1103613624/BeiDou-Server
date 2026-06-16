from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ---------- 配置信息 ----------
sheets_info = {
    "目录导航": {
        "用途": "全表索引入口",
        "适用人群": "所有玩家",
        "板块": ["分表快速跳转链接", "版本更新说明", "常用术语解释", "攻略更新日志"]
    },
    "新手入门攻略": {
        "用途": "新手玩家引导",
        "适用人群": "刚接触GMS083的新玩家",
        "板块": ["职业选择推荐", "1-200级快速升级路线", "基础操作指南", "初期资源获取攻略"]
    },
    "职业进阶攻略": {
        "用途": "各职业深度养成",
        "适用人群": "有基础想深耕职业的玩家",
        "板块": ["全职业技能加点方案", "各职业装备搭配推荐", "职业核心技能循环教学", "各职业PVE/PVP强度解析"]
    },
    "副本通关攻略": {
        "用途": "副本玩法指南",
        "适用人群": "需要刷副本获取资源的玩家",
        "板块": ["普通/困难/极限副本通关流程", "副本BOSS机制详解", "副本掉落清单及概率", "副本组队配置推荐"]
    },
    "资源收集攻略": {
        "用途": "全资源获取指引",
        "适用人群": "需要积累游戏资源的玩家",
        "板块": ["金币/枫币快速获取方法", "稀有道具掉落点位", "生活技能收益最大化指南", "活动资源兑换优先级"]
    },
    "活动指南": {
        "用途": "版本活动说明",
        "适用人群": "想参与版本活动的玩家",
        "板块": ["当期活动开启时间", "活动任务流程", "活动奖励清单", "活动隐藏福利说明"]
    },
    "玩家意见收集": {
        "用途": "反馈与建议汇总",
        "适用人群": "所有想反馈意见/提需求的玩家",
        "板块": ["攻略内容纠错反馈入口", "玩家玩法建议收集", "游戏BUG反馈通道", "攻略需求征集"]
    }
}

# ---------- 创建工作簿 ----------
wb = Workbook()
wb.remove(wb.active)  # 删除默认的 Sheet

# 先创建所有工作表（保证目录导航超链接有效）
for name in sheets_info.keys():
    wb.create_sheet(name)

# ---------- 填充各内容工作表 ----------
for name, info in sheets_info.items():
    if name == "目录导航":
        continue
    ws = wb[name]
    # 标题区：用途、适用人群
    ws['A1'] = "用途："
    ws['B1'] = info["用途"]
    ws['A2'] = "适用人群："
    ws['B2'] = info["适用人群"]
    # 加粗标题
    for cell in ['A1','B1','A2','B2']:
        ws[cell].font = Font(bold=True)
    # 板块列表
    ws['A4'] = "板块名称"
    ws['B4'] = "具体内容 / 说明"
    ws['C4'] = "备注"
    for r, heading in enumerate(info["板块"], start=5):
        ws[f'A{r}'] = heading
        ws[f'B{r}'] = "待补充"
        ws[f'C{r}'] = ""
    # 列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20

# ---------- 填充目录导航 ----------
ws_dir = wb["目录导航"]
# 表头
headers = ["序号", "工作表名称", "用途", "适用人群", "跳转链接"]
for col_idx, h in enumerate(headers, 1):
    cell = ws_dir.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True)
# 填充各表信息
row = 2
for name, info in sheets_info.items():
    if name == "目录导航":
        continue
    ws_dir.cell(row=row, column=1, value=row-1)
    ws_dir.cell(row=row, column=2, value=name)
    ws_dir.cell(row=row, column=3, value=info["用途"])
    ws_dir.cell(row=row, column=4, value=info["适用人群"])
    # 超链接公式
    ws_dir.cell(row=row, column=5, value=f'=HYPERLINK("#\'{name}\'!A1", "跳转")')
    row += 1

# 右侧附加信息：版本、术语、日志
ws_dir['G1'] = "版本更新说明"
ws_dir['G1'].font = Font(bold=True)
ws_dir['G2'] = "v1.0 初始模板"

ws_dir['H1'] = "常用术语解释"
ws_dir['H1'].font = Font(bold=True)
ws_dir['H2'] = "待补充"

ws_dir['I1'] = "攻略更新日志"
ws_dir['I1'].font = Font(bold=True)
ws_dir['I2'] = "2026-06-16 创建"

# 设置列宽
for col in ['A','B','C','D','E','G','H','I']:
    ws_dir.column_dimensions[col].width = 20

# 保存文件
wb.save("GMS083攻略表格模板.xlsx")
print("✅ Excel 模板已生成：GMS083攻略表格模板.xlsx")